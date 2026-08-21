#!/usr/bin/env python3
"""
Builds SavePlate_Test_Report.xlsx: an Excel workbook matching the course's
Unit Testing / System Testing templates, populated with real test cases from
tests/uc1-register-login-2fa.spec.js and screenshots captured from an actual
run of the app. System testing uses Playwright in place of Selenium IDE.

One-off report generator, not part of the app itself -- run manually:
    python3 scripts/build-test-report-xlsx.py
"""
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

ROOT = "/Users/shekharbhatt/Desktop/saveplate"
EVIDENCE = f"{ROOT}/test-evidence"
OUT_PATH = f"{ROOT}/SavePlate_Test_Report.xlsx"

TODAY = datetime.date.today().strftime("%Y-%m-%d")

YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def style_label_value_row(ws, row, label, value, highlight=False, value_font=None):
    a = ws.cell(row=row, column=1, value=label)
    b = ws.cell(row=row, column=2, value=value)
    a.font = BOLD
    a.border = BORDER
    a.alignment = WRAP
    b.border = BORDER
    b.alignment = WRAP
    if value_font:
        b.font = value_font
    if highlight:
        b.fill = YELLOW
    return row + 1


def add_image(ws, path, anchor_cell, width_px=460):
    img = XLImage(path)
    scale = width_px / img.width
    img.width = width_px
    img.height = int(img.height * scale)
    img.anchor = anchor_cell
    ws.add_image(img)
    # Rows needed so the sheet visually reserves space for the image.
    return max(2, int(img.height / 15) + 1)


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = Workbook()

# ============================================================
# SHEET 1: Unit Testing (individual)
# ============================================================
ws1 = wb.active
ws1.title = "Unit Testing"
set_widths(ws1, [26, 70, 4])

row = 1
title = ws1.cell(row=row, column=1, value="Unit Testing (individual)")
title.font = Font(bold=True, size=13)
row += 2

unit_cases = [
    {
        "tc_id": "TC1",
        "scenario": "Customer Login-in with valid data",
        "test_data": "Email: demo@saveplate.com\nPassword: demo123",
        "source_code": (
            "const user = users.find(u =>\n"
            "  u.email.toLowerCase() === emailEl.value.trim().toLowerCase()\n"
            "  && u.password === passEl.value);\n\n"
            "if (!user) {\n"
            "  // show \"Incorrect password or email.\"\n"
            "} else {\n"
            "  db.setCurrentUser(user);\n"
            "  window.location.hash = '#dashboard';\n"
            "  showView('dashboard');\n"
            "}"
        ),
        "expected": "User is authenticated and redirected to the Dashboard (#dashboard).",
        "screenshot": f"{EVIDENCE}/unit_TC1_valid_login.png",
        "pass_fail": "PASS",
        "remarks": "Automated via Playwright (tests/uc1-register-login-2fa.spec.js:53).",
    },
    {
        "tc_id": "TC2",
        "scenario": "Customer Login-in with invalid UserID and Password",
        "test_data": "Email: demo@saveplate.com\nPassword: wrongpassword",
        "source_code": (
            "const user = users.find(u =>\n"
            "  u.email.toLowerCase() === emailEl.value.trim().toLowerCase()\n"
            "  && u.password === passEl.value);\n\n"
            "if (!user) {\n"
            "  passEl.parentElement.classList.add('invalid');\n"
            "  errorMsg.textContent = 'Incorrect password or email.';\n"
            "  toast('Login failed. Please check your credentials.', 'error');\n"
            "  return;\n"
            "}"
        ),
        "expected": "Login is rejected; inline error \"Incorrect password or email.\" and an error toast are shown. User stays on the login screen.",
        "screenshot": f"{EVIDENCE}/unit_TC2_invalid_login.png",
        "pass_fail": "PASS",
        "remarks": "Automated via Playwright (tests/uc1-register-login-2fa.spec.js:77).",
    },
]

for case in unit_cases:
    row = style_label_value_row(ws1, row, "Tester Name", "")
    row = style_label_value_row(ws1, row, "Date", TODAY)
    row = style_label_value_row(ws1, row, "Test Description", "Verify username/password")
    row = style_label_value_row(ws1, row, "Test Case ID", case["tc_id"])
    row = style_label_value_row(ws1, row, "Test Scenario", case["scenario"])
    row = style_label_value_row(ws1, row, "Test Data", case["test_data"], highlight=True)
    row = style_label_value_row(
        ws1, row, "Source Code (javascript/php)", case["source_code"],
        value_font=Font(name="Consolas", size=9)
    )
    ws1.row_dimensions[row - 1].height = 110
    row = style_label_value_row(ws1, row, "Expected Output", case["expected"], highlight=True)

    actual_label_row = row
    a = ws1.cell(row=actual_label_row, column=1, value="Actual Output")
    a.font = BOLD
    a.border = BORDER
    b = ws1.cell(row=actual_label_row, column=2, value="Print Screen (below)")
    b.fill = YELLOW
    b.border = BORDER
    row += 1
    img_rows = add_image(ws1, case["screenshot"], f"B{row}", width_px=460)
    row += img_rows + 1

    row = style_label_value_row(ws1, row, "Pass/Fail", case["pass_fail"])
    row = style_label_value_row(ws1, row, "Remarks", case["remarks"])
    row += 2  # spacer between test case blocks

row1 = row  # next free row on ws1, resumed later for the UC3 unit-test block

# ============================================================
# SHEET 2: System Testing (Playwright, replacing Selenium IDE)
# ============================================================
ws2 = wb.create_sheet("System Testing")
set_widths(ws2, [26, 70, 4])

row = 1
title2 = ws2.cell(
    row=row, column=1,
    value="System Testing (use Playwright) — replaces Selenium IDE; spec file: tests/uc1-register-login-2fa.spec.js"
)
title2.font = Font(bold=True, size=12)
ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
ws2.cell(row=row, column=1).alignment = WRAP
row += 2

row = style_label_value_row(ws2, row, "Functional Requirement", "Sign-up")
row += 1

system_cases = [
    {
        "tc_id": "TC1",
        "scenario": "Sign-up with invalid data (empty name, bad email, out-of-range household size, short/mismatched password)",
        "test_data": "Name: (empty)\nEmail: not-an-email\nHousehold Size: 99\nPassword: 123\nConfirm Password: 456",
        "steps": (
            "1. Go to the app, click \"Register here\".\n"
            "2. Leave Full Name empty.\n"
            "3. Enter an invalid email, size 99, password '123', confirm '456'.\n"
            "4. Click \"Register Account\"."
        ),
        "expected": "All five fields are flagged invalid (red outline + inline error text); no account is created; user stays on the Create Account screen.",
        "screenshot": f"{EVIDENCE}/system_TC1_signup_validation_errors.png",
        "pass_fail": "PASS",
        "remarks": "tests/uc1-register-login-2fa.spec.js:17",
    },
    {
        "tc_id": "TC2",
        "scenario": "Sign-up with an email that is already registered",
        "test_data": "Name: Duplicate Tester\nEmail: demo@saveplate.com (existing)\nPassword/Confirm: validpass",
        "steps": (
            "1. Click \"Register here\".\n"
            "2. Fill in a valid name/password using the already-registered demo email.\n"
            "3. Click \"Register Account\"."
        ),
        "expected": "Email field is flagged invalid with the message \"This email is already registered.\"; no duplicate account is created.",
        "screenshot": f"{EVIDENCE}/system_TC2_signup_duplicate_email.png",
        "pass_fail": "PASS",
        "remarks": "tests/uc1-register-login-2fa.spec.js:38",
    },
    {
        "tc_id": "TC3",
        "scenario": "Sign-up successfully with valid data (no 2FA)",
        "test_data": "Name: Evidence Tester\nEmail: evidence_<timestamp>@saveplate.com\nPassword/Confirm: demo123",
        "steps": (
            "1. Click \"Register here\".\n"
            "2. Fill in valid name, a unique email, matching password/confirm.\n"
            "3. Leave \"Enable Two-Factor Authentication\" unchecked.\n"
            "4. Click \"Register Account\"."
        ),
        "expected": "Account is created; user is returned to the Login screen with their email pre-filled and a success toast shown.",
        "screenshot": f"{EVIDENCE}/system_TC3_signup_success.png",
        "pass_fail": "PASS",
        "remarks": "tests/uc1-register-login-2fa.spec.js:53",
    },
    {
        "tc_id": "TC4",
        "scenario": "Sign-up with 2FA enabled, then log in using the generated authenticator code",
        "test_data": "Name: Evidence 2FA Tester\nEmail: evidence2fa_<timestamp>@saveplate.com\nPassword/Confirm: demo123\n2FA: enabled",
        "steps": (
            "1. Click \"Register here\", fill valid details, check \"Enable Two-Factor Authentication (2FA)\".\n"
            "2. Click \"Register Account\" → a \"Set Up 2FA\" modal opens with a live 6-digit code.\n"
            "3. Click \"Confirm\" to complete setup.\n"
            "4. Log in with the new account's email/password.\n"
            "5. On the 2FA Verification screen, submit the live code shown."
        ),
        "expected": "2FA secret is saved; login is blocked until a valid TOTP code is entered; correct code logs the user into the Dashboard.",
        "screenshot": f"{EVIDENCE}/system_TC4_signup_2fa_setup.png",
        "pass_fail": "PASS",
        "remarks": "tests/uc1-register-login-2fa.spec.js:90",
    },
]

def add_system_case(ws, row, case):
    row = style_label_value_row(ws, row, "Test Case ID", case["tc_id"])
    row = style_label_value_row(ws, row, "Test Scenario", case["scenario"])
    row = style_label_value_row(ws, row, "Test Data", case["test_data"], highlight=True)
    row = style_label_value_row(ws, row, "Test Steps", case["steps"])
    ws.row_dimensions[row - 1].height = 80
    row = style_label_value_row(ws, row, "Expected Output", case["expected"], highlight=True)

    a = ws.cell(row=row, column=1, value="Actual Output (Screen Shots)")
    a.font = BOLD
    a.border = BORDER
    b = ws.cell(row=row, column=2, value="See screenshot below")
    b.fill = YELLOW
    b.border = BORDER
    row += 1
    screenshots = case.get("screenshots") or [case["screenshot"]]
    for shot in screenshots:
        img_rows = add_image(ws, shot, f"B{row}", width_px=460)
        row += img_rows + 1

    row = style_label_value_row(ws, row, "Pass/Fail (log)", case["pass_fail"])
    row = style_label_value_row(ws, row, "Remarks", case["remarks"])
    row += 2
    return row


for case in system_cases:
    row = add_system_case(ws2, row, case)

# Overall automation log screenshot (Selenium Log -> Playwright HTML report)
row = style_label_value_row(
    ws2, row, "Screen Shot of Playwright Log",
    "Full UC1 suite run (7/7 passed) — replaces \"Screen Shot of Selenium Log\""
)
img_rows = add_image(ws2, f"{ROOT}/uc1-playwright-report.png", f"B{row}", width_px=620)
row += img_rows + 2

# ============================================================
# UC3: Browse & Donate — Unit Testing (individual)
# ============================================================
row1 = style_label_value_row(
    ws1, row1, "Section",
    "UC3: Browse & Donate (browse.js) — spec file: tests/uc3-browse-donate.spec.js"
)
row1 += 1

uc3_unit_cases = [
    {
        "tc_id": "TC1",
        "scenario": "Search filter narrows the donation grid by item name",
        "test_data": "Search text: bagels",
        "source_code": (
            "const searchQuery = document.getElementById('browse-search')\n"
            "  .value.toLowerCase().trim();\n\n"
            "if (searchQuery) {\n"
            "  donations = donations.filter(d =>\n"
            "    d.name.toLowerCase().includes(searchQuery));\n"
            "}"
        ),
        "expected": "Only donations whose name contains \"bagels\" remain in the grid; the totals label updates to \"Showing 1 community donation available\".",
        "screenshot": f"{EVIDENCE}/system_TC2_uc3_search_filter.png",
        "pass_fail": "PASS",
        "remarks": "Automated via Playwright (tests/uc3-browse-donate.spec.js:30).",
    },
    {
        "tc_id": "TC2",
        "scenario": "Prevent a user from claiming their own donation",
        "test_data": "Donation donorId === current user's id (e.g. \"Canned Lentils\")",
        "source_code": (
            "if (item.donorId === user.id) {\n"
            "  toast('You cannot claim your own donation!', 'error');\n"
            "  return;\n"
            "}\n\n"
            "// Card render: own donation gets a disabled button\n"
            "isOwnDonation ?\n"
            "  `<button ... disabled>Your Item</button>` :\n"
            "  `<button ... onclick=\"claimDonatedItem('${item.id}')\">Claim Item</button>`"
        ),
        "expected": "The donor's own item shows a disabled \"Your Item\" button (no \"Claim Item\" control); the donor field reads \"You (My Donation)\".",
        "screenshot": f"{EVIDENCE}/system_TC5_uc3_own_donation_disabled.png",
        "pass_fail": "PASS",
        "remarks": "Automated via Playwright (tests/uc3-browse-donate.spec.js:62).",
    },
]

for case in uc3_unit_cases:
    row1 = style_label_value_row(ws1, row1, "Tester Name", "")
    row1 = style_label_value_row(ws1, row1, "Date", TODAY)
    row1 = style_label_value_row(ws1, row1, "Test Description", "Verify community donation search/filter and claim-ownership rules")
    row1 = style_label_value_row(ws1, row1, "Test Case ID", case["tc_id"])
    row1 = style_label_value_row(ws1, row1, "Test Scenario", case["scenario"])
    row1 = style_label_value_row(ws1, row1, "Test Data", case["test_data"], highlight=True)
    row1 = style_label_value_row(
        ws1, row1, "Source Code (javascript/php)", case["source_code"],
        value_font=Font(name="Consolas", size=9)
    )
    ws1.row_dimensions[row1 - 1].height = 110
    row1 = style_label_value_row(ws1, row1, "Expected Output", case["expected"], highlight=True)

    actual_label_row = row1
    a = ws1.cell(row=actual_label_row, column=1, value="Actual Output")
    a.font = BOLD
    a.border = BORDER
    b = ws1.cell(row=actual_label_row, column=2, value="Print Screen (below)")
    b.fill = YELLOW
    b.border = BORDER
    row1 += 1
    img_rows = add_image(ws1, case["screenshot"], f"B{row1}", width_px=460)
    row1 += img_rows + 1

    row1 = style_label_value_row(ws1, row1, "Pass/Fail", case["pass_fail"])
    row1 = style_label_value_row(ws1, row1, "Remarks", case["remarks"])
    row1 += 2

# ============================================================
# UC3: Browse & Donate — System Testing (Playwright)
# ============================================================
row2 = row
row2 = style_label_value_row(
    ws2, row2, "Section",
    "UC3: Browse & Donate (use Playwright) — spec file: tests/uc3-browse-donate.spec.js"
)
row2 += 1
row2 = style_label_value_row(ws2, row2, "Functional Requirement", "Browse & Donate")
row2 += 1

uc3_system_cases = [
    {
        "tc_id": "TC1",
        "scenario": "Renders seeded community donations with donor and details",
        "test_data": "Logged in as demo@saveplate.com (seeded dataset: Bagels, Organic Strawberries, Sweet Corn Cans)",
        "steps": (
            "1. Log in as the demo account.\n"
            "2. Click \"Browse & Donate\" in the sidebar."
        ),
        "expected": "Grid shows 3 donation cards (\"Showing 3 community donations available\"); each card shows name, quantity, category, storage, expiry, and donor (e.g. \"Sarah Jenkins\" on Organic Strawberries).",
        "screenshot": f"{EVIDENCE}/system_TC1_uc3_browse_grid.png",
        "pass_fail": "PASS",
        "remarks": "tests/uc3-browse-donate.spec.js:20",
    },
    {
        "tc_id": "TC2",
        "scenario": "Search narrows the donation grid by item name",
        "test_data": "Search text: bagels",
        "steps": (
            "1. On Browse & Donate, type \"bagels\" into the search box."
        ),
        "expected": "Grid narrows to a single card, \"Bagels (Sesame)\"; totals label reads \"Showing 1 community donation available\".",
        "screenshot": f"{EVIDENCE}/system_TC2_uc3_search_filter.png",
        "pass_fail": "PASS",
        "remarks": "tests/uc3-browse-donate.spec.js:30",
    },
    {
        "tc_id": "TC3",
        "scenario": "Category filter narrows the donation grid",
        "test_data": "Category filter: Canned",
        "steps": (
            "1. On Browse & Donate, select \"Canned\" from the category dropdown."
        ),
        "expected": "Grid narrows to a single card, \"Sweet Corn Cans\"; totals label reads \"Showing 1 community donation available\".",
        "screenshot": f"{EVIDENCE}/system_TC3_uc3_category_filter.png",
        "pass_fail": "PASS",
        "remarks": "tests/uc3-browse-donate.spec.js:39",
    },
    {
        "tc_id": "TC4",
        "scenario": "Claiming a donation removes it from the board and notifies the claimer",
        "test_data": "Claim target: Bagels (Sesame), donated by Clara Oswald",
        "steps": (
            "1. On \"Bagels (Sesame)\", click \"Claim Item\".\n"
            "2. Click \"Confirm\" in the claim dialog.\n"
            "3. Open \"Notifications\" in the sidebar."
        ),
        "expected": "The card disappears from the board (\"Showing 2 community donations available\"); a \"Donation Claimed: Bagels (Sesame)\" notification appears for the claimer.",
        "screenshots": [
            f"{EVIDENCE}/system_TC4a_uc3_claim_confirmed.png",
            f"{EVIDENCE}/system_TC4b_uc3_claim_notification.png",
        ],
        "pass_fail": "PASS",
        "remarks": "tests/uc3-browse-donate.spec.js:48",
    },
    {
        "tc_id": "TC5",
        "scenario": "Cannot claim your own donation",
        "test_data": "Donate \"Canned Lentils\" from inventory, then view it on the community board",
        "steps": (
            "1. On \"Food Inventory\", click \"Donate to Community Board\" for \"Canned Lentils\" and confirm.\n"
            "2. Go to \"Browse & Donate\"."
        ),
        "expected": "The \"Canned Lentils\" card shows donor \"You (My Donation)\" and a disabled \"Your Item\" button; no \"Claim Item\" button is shown.",
        "screenshot": f"{EVIDENCE}/system_TC5_uc3_own_donation_disabled.png",
        "pass_fail": "PASS",
        "remarks": "tests/uc3-browse-donate.spec.js:62",
    },
]

for case in uc3_system_cases:
    row2 = add_system_case(ws2, row2, case)

row2 = style_label_value_row(
    ws2, row2, "Screen Shot of Playwright Log",
    "Full UC3 suite run (5/5 passed) — replaces \"Screen Shot of Selenium Log\""
)
img_rows = add_image(ws2, f"{ROOT}/uc3-playwright-report.png", f"B{row2}", width_px=620)
row2 += img_rows + 2

# ============================================================
# UC2: Food Inventory — Unit Testing (individual)
# ============================================================
row1 = style_label_value_row(
    ws1, row1, "Section",
    "UC2: Food Inventory (inventory.js) — spec file: tests/uc2-inventory.spec.js"
)
row1 += 1

uc2_unit_cases = [
    {
        "tc_id": "TC1",
        "scenario": "Expiry status pill classification (Expired / Use in Xd / Fresh)",
        "test_data": "Whole Wheat Bread: expired 2026-07-21\nRoma Tomatoes: expires today\nFuji Apples: expires 2026-08-03",
        "source_code": (
            "const days = daysUntil(item.expiryDate);\n"
            "if (days < 0) {\n"
            "  statusHtml = `<span class=\"pill pill-danger\">Expired</span>`;\n"
            "} else if (days <= 2) {\n"
            "  statusHtml = `<span class=\"pill pill-warning\">Use in ${days}d</span>`;\n"
            "} else {\n"
            "  statusHtml = `<span class=\"pill pill-fresh\">Fresh</span>`;\n"
            "}"
        ),
        "expected": "Whole Wheat Bread shows a red \"Expired\" pill, Roma Tomatoes shows an amber \"Use in 0d\" pill, and Fuji Apples shows a green \"Fresh\" pill.",
        "screenshot": f"{EVIDENCE}/system_TC1_uc2_status_pills.png",
        "pass_fail": "PASS",
        "remarks": "Automated via Playwright (tests/uc2-inventory.spec.js:31).",
    },
    {
        "tc_id": "TC2",
        "scenario": "Add Food Item form validation rejects empty name, zero quantity, blank expiry",
        "test_data": "Food Name: (empty)\nQuantity: 0\nExpiry Date: (empty)",
        "source_code": (
            "if (!nameEl.value.trim()) {\n"
            "  nameEl.parentElement.classList.add('invalid'); isValid = false;\n"
            "}\n"
            "const qty = parseFloat(qtyEl.value);\n"
            "if (isNaN(qty) || qty <= 0) {\n"
            "  qtyEl.parentElement.classList.add('invalid'); isValid = false;\n"
            "}\n"
            "if (!expiryEl.value) {\n"
            "  expiryEl.parentElement.classList.add('invalid'); isValid = false;\n"
            "}\n"
            "if (!isValid) return;"
        ),
        "expected": "Name and Quantity fields are flagged invalid with inline error text; the modal stays open and no item is saved.",
        "screenshot": f"{EVIDENCE}/system_TC4_uc2_add_validation_errors.png",
        "pass_fail": "PASS",
        "remarks": "Automated via Playwright (tests/uc2-inventory.spec.js:51).",
    },
]

for case in uc2_unit_cases:
    row1 = style_label_value_row(ws1, row1, "Tester Name", "")
    row1 = style_label_value_row(ws1, row1, "Date", TODAY)
    row1 = style_label_value_row(ws1, row1, "Test Description", "Verify inventory expiry-status logic and add-item validation")
    row1 = style_label_value_row(ws1, row1, "Test Case ID", case["tc_id"])
    row1 = style_label_value_row(ws1, row1, "Test Scenario", case["scenario"])
    row1 = style_label_value_row(ws1, row1, "Test Data", case["test_data"], highlight=True)
    row1 = style_label_value_row(
        ws1, row1, "Source Code (javascript/php)", case["source_code"],
        value_font=Font(name="Consolas", size=9)
    )
    ws1.row_dimensions[row1 - 1].height = 110
    row1 = style_label_value_row(ws1, row1, "Expected Output", case["expected"], highlight=True)

    actual_label_row = row1
    a = ws1.cell(row=actual_label_row, column=1, value="Actual Output")
    a.font = BOLD
    a.border = BORDER
    b = ws1.cell(row=actual_label_row, column=2, value="Print Screen (below)")
    b.fill = YELLOW
    b.border = BORDER
    row1 += 1
    img_rows = add_image(ws1, case["screenshot"], f"B{row1}", width_px=460)
    row1 += img_rows + 1

    row1 = style_label_value_row(ws1, row1, "Pass/Fail", case["pass_fail"])
    row1 = style_label_value_row(ws1, row1, "Remarks", case["remarks"])
    row1 += 2

# ============================================================
# UC2: Food Inventory — System Testing (Playwright)
# ============================================================
row2 = style_label_value_row(
    ws2, row2, "Section",
    "UC2: Food Inventory (use Playwright) — spec file: tests/uc2-inventory.spec.js"
)
row2 += 1
row2 = style_label_value_row(ws2, row2, "Functional Requirement", "Food Inventory Management")
row2 += 1

uc2_system_cases = [
    {
        "tc_id": "TC1",
        "scenario": "Shows correct status pills for expired, expiring-soon, and fresh seeded items",
        "test_data": "Seeded items: Whole Wheat Bread (expired), Roma Tomatoes (expires today), Fuji Apples (expires 2026-08-03)",
        "steps": (
            "1. Log in as the demo account.\n"
            "2. Click \"Food Inventory\" in the sidebar."
        ),
        "expected": "Whole Wheat Bread shows a red \"Expired\" pill; Roma Tomatoes shows an amber \"Use in 0d\" pill; Fuji Apples shows a green \"Fresh\" pill.",
        "screenshot": f"{EVIDENCE}/system_TC1_uc2_status_pills.png",
        "pass_fail": "PASS",
        "remarks": "tests/uc2-inventory.spec.js:31",
    },
    {
        "tc_id": "TC2",
        "scenario": "Search narrows the list by item name",
        "test_data": "Search text: avocado",
        "steps": (
            "1. On Food Inventory, type \"avocado\" into the search box."
        ),
        "expected": "The table narrows to a single row, \"Ripe Avocado\".",
        "screenshot": f"{EVIDENCE}/system_TC2_uc2_search_filter.png",
        "pass_fail": "PASS",
        "remarks": "tests/uc2-inventory.spec.js:37",
    },
    {
        "tc_id": "TC3",
        "scenario": "Category filter narrows the list",
        "test_data": "Category filter: Dairy",
        "steps": (
            "1. On Food Inventory, select \"Dairy\" from the category dropdown."
        ),
        "expected": "The table narrows to a single row, \"Whole Milk\".",
        "screenshot": f"{EVIDENCE}/system_TC3_uc2_category_filter.png",
        "pass_fail": "PASS",
        "remarks": "tests/uc2-inventory.spec.js:44",
    },
    {
        "tc_id": "TC4",
        "scenario": "Rejects an empty name, zero quantity, and blank expiry date",
        "test_data": "Food Name: (empty)\nQuantity: 0\nExpiry Date: (empty)",
        "steps": (
            "1. Click \"Add Food Item\".\n"
            "2. Set Quantity to \"0\", leave Food Name and Expiry Date empty.\n"
            "3. Click \"Confirm\"."
        ),
        "expected": "Food Name and Quantity fields are flagged invalid with inline error text; the modal stays open and no item is saved.",
        "screenshot": f"{EVIDENCE}/system_TC4_uc2_add_validation_errors.png",
        "pass_fail": "PASS",
        "remarks": "tests/uc2-inventory.spec.js:51",
    },
    {
        "tc_id": "TC5",
        "scenario": "Adds a new food item and displays it in the table",
        "test_data": "Name: Greek Yogurt\nQuantity: 2 tubs\nCategory: Dairy\nLocation: Fridge\nExpiry: 2027-01-01",
        "steps": (
            "1. Click \"Add Food Item\" and fill in valid details.\n"
            "2. Click \"Confirm\"."
        ),
        "expected": "The modal closes and \"Greek Yogurt\" appears as a new row in the inventory table.",
        "screenshot": f"{EVIDENCE}/system_TC5_uc2_add_item_success.png",
        "pass_fail": "PASS",
        "remarks": "tests/uc2-inventory.spec.js:64",
    },
    {
        "tc_id": "TC6",
        "scenario": "Marking an item used removes it from the active inventory list",
        "test_data": "Item: Jasmine Rice",
        "steps": (
            "1. Click the \"Mark Used / Eaten\" icon on the \"Jasmine Rice\" row."
        ),
        "expected": "The \"Jasmine Rice\" row is removed from the active inventory table.",
        "screenshot": f"{EVIDENCE}/system_TC6_uc2_mark_used.png",
        "pass_fail": "PASS",
        "remarks": "tests/uc2-inventory.spec.js:89",
    },
    {
        "tc_id": "TC7",
        "scenario": "Donating an item removes it from the personal inventory list",
        "test_data": "Item: Canned Lentils",
        "steps": (
            "1. Click \"Donate to Community Board\" on the \"Canned Lentils\" row.\n"
            "2. Click \"Confirm\"."
        ),
        "expected": "The \"Canned Lentils\" row is removed from the personal inventory table (it now appears on the community board — see UC3).",
        "screenshot": f"{EVIDENCE}/system_TC7_uc2_donate_item.png",
        "pass_fail": "PASS",
        "remarks": "tests/uc2-inventory.spec.js:95",
    },
    {
        "tc_id": "TC8",
        "scenario": "Deletes an item after confirmation",
        "test_data": "Item: Chicken Breasts",
        "steps": (
            "1. Click the \"Delete Item\" icon on the \"Chicken Breasts\" row.\n"
            "2. Click \"Confirm\" in the delete dialog."
        ),
        "expected": "A confirmation dialog appears (\"Are you sure you want to delete this item?\"); after confirming, the \"Chicken Breasts\" row is removed.",
        "screenshots": [
            f"{EVIDENCE}/system_TC8a_uc2_delete_confirm_dialog.png",
            f"{EVIDENCE}/system_TC8b_uc2_delete_success.png",
        ],
        "pass_fail": "PASS",
        "remarks": "tests/uc2-inventory.spec.js:102",
    },
]

for case in uc2_system_cases:
    row2 = add_system_case(ws2, row2, case)

row2 = style_label_value_row(
    ws2, row2, "Screen Shot of Playwright Log",
    "Full UC2 suite run (9/9 passed) — replaces \"Screen Shot of Selenium Log\""
)
add_image(ws2, f"{ROOT}/uc2-playwright-report.png", f"B{row2}", width_px=620)

wb.save(OUT_PATH)
print("Wrote", OUT_PATH)
